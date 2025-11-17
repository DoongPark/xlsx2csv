#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
테스트용 xlsx 파일 생성 도구

다양한 크기의 xlsx 파일을 생성하여 성능 테스트에 사용
"""

import os
import sys
import random
import string
from datetime import datetime, timedelta
import argparse

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("openpyxl 패키지가 필요합니다. 설치하려면: pip install openpyxl")
    sys.exit(1)


class TestDataGenerator:
    """테스트 데이터 생성기"""
    
    def __init__(self, output_dir="test_data"):
        self.output_dir = output_dir
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
    
    def random_string(self, length=10):
        """랜덤 문자열 생성"""
        return ''.join(random.choices(string.ascii_letters + string.digits, k=length))
    
    def random_date(self, start_year=2020, end_year=2024):
        """랜덤 날짜 생성"""
        start = datetime(start_year, 1, 1)
        end = datetime(end_year, 12, 31)
        delta = end - start
        random_days = random.randint(0, delta.days)
        return start + timedelta(days=random_days)
    
    def generate_row_data(self, num_columns):
        """단일 행 데이터 생성"""
        row = []
        for i in range(num_columns):
            data_type = random.choice(['string', 'integer', 'float', 'date', 'boolean'])
            
            if data_type == 'string':
                row.append(self.random_string(random.randint(5, 30)))
            elif data_type == 'integer':
                row.append(random.randint(0, 1000000))
            elif data_type == 'float':
                row.append(round(random.uniform(0, 10000), 2))
            elif data_type == 'date':
                row.append(self.random_date())
            elif data_type == 'boolean':
                row.append(random.choice([True, False]))
        
        return row
    
    def create_xlsx_file(self, filename, num_sheets, rows_per_sheet, columns_per_sheet, 
                        add_formatting=False, verbose=True):
        """
        xlsx 파일 생성
        
        Args:
            filename: 출력 파일명
            num_sheets: 시트 개수
            rows_per_sheet: 시트당 행 수
            columns_per_sheet: 시트당 열 수
            add_formatting: 서식 추가 여부
            verbose: 진행상황 출력 여부
        """
        filepath = os.path.join(self.output_dir, filename)
        
        if verbose:
            print(f"\n생성 중: {filename}")
            print(f"  시트 수: {num_sheets}")
            print(f"  시트당 행 수: {rows_per_sheet:,}")
            print(f"  시트당 열 수: {columns_per_sheet}")
        
        wb = Workbook()
        # 기본 시트 제거
        wb.remove(wb.active)
        
        for sheet_num in range(1, num_sheets + 1):
            if verbose:
                print(f"  시트 {sheet_num}/{num_sheets} 생성 중...", end='', flush=True)
            
            ws = wb.create_sheet(f"Sheet{sheet_num}")
            
            # 헤더 추가
            header = [f"Column{i+1}" for i in range(columns_per_sheet)]
            ws.append(header)
            
            # 헤더 서식
            if add_formatting:
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", 
                                           end_color="CCCCCC", 
                                           fill_type="solid")
            
            # 데이터 행 추가
            batch_size = 1000
            for batch_start in range(0, rows_per_sheet, batch_size):
                batch_end = min(batch_start + batch_size, rows_per_sheet)
                for _ in range(batch_end - batch_start):
                    ws.append(self.generate_row_data(columns_per_sheet))
            
            if verbose:
                print(" 완료")
        
        # 파일 저장
        wb.save(filepath)
        file_size = os.path.getsize(filepath) / (1024 * 1024)  # MB
        
        if verbose:
            print(f"✓ 생성 완료: {filepath} ({file_size:.2f} MB)")
        
        return filepath
    
    def generate_test_suite(self):
        """표준 테스트 스위트 생성"""
        print("=" * 60)
        print("xlsx2csv 성능 테스트 데이터 생성")
        print("=" * 60)
        
        test_configs = [
            # (파일명, 시트수, 행수, 열수, 설명)
            ("small_multi_sheet.xlsx", 10, 1000, 10, "소형: 다중 시트"),
            ("small_single_sheet.xlsx", 1, 10000, 10, "소형: 단일 시트"),
            
            ("medium_multi_sheet.xlsx", 5, 10000, 20, "중형: 다중 시트"),
            ("medium_single_sheet.xlsx", 1, 50000, 20, "중형: 단일 시트"),
            
            ("large_multi_sheet.xlsx", 10, 100000, 50, "대형: 다중 시트"),
            ("large_single_sheet.xlsx", 1, 500000, 50, "대형: 단일 시트"),
            
            ("xlarge_multi_sheet.xlsx", 20, 50000, 30, "초대형: 다중 시트"),
            ("xlarge_single_sheet.xlsx", 1, 1000000, 100, "초대형: 단일 시트"),
        ]
        
        results = []
        
        for filename, sheets, rows, cols, description in test_configs:
            print(f"\n[{description}]")
            try:
                filepath = self.create_xlsx_file(
                    filename=filename,
                    num_sheets=sheets,
                    rows_per_sheet=rows,
                    columns_per_sheet=cols,
                    add_formatting=False,
                    verbose=True
                )
                results.append((filename, "성공", description))
            except Exception as e:
                print(f"✗ 실패: {str(e)}")
                results.append((filename, f"실패: {str(e)}", description))
        
        # 결과 요약
        print("\n" + "=" * 60)
        print("생성 결과 요약")
        print("=" * 60)
        for filename, status, description in results:
            print(f"{filename:30s} - {status}")
        
        print(f"\n모든 파일이 '{self.output_dir}' 디렉토리에 저장되었습니다.")


def main():
    parser = argparse.ArgumentParser(
        description="xlsx2csv 성능 테스트용 xlsx 파일 생성"
    )
    parser.add_argument(
        '--output-dir', 
        default='test_data',
        help='출력 디렉토리 (기본값: test_data)'
    )
    parser.add_argument(
        '--custom',
        action='store_true',
        help='사용자 정의 파일 생성 모드'
    )
    parser.add_argument(
        '--filename',
        default='custom.xlsx',
        help='파일명 (custom 모드용)'
    )
    parser.add_argument(
        '--sheets',
        type=int,
        default=5,
        help='시트 수 (custom 모드용, 기본값: 5)'
    )
    parser.add_argument(
        '--rows',
        type=int,
        default=10000,
        help='시트당 행 수 (custom 모드용, 기본값: 10000)'
    )
    parser.add_argument(
        '--columns',
        type=int,
        default=20,
        help='시트당 열 수 (custom 모드용, 기본값: 20)'
    )
    parser.add_argument(
        '--with-formatting',
        action='store_true',
        help='서식 추가'
    )
    
    args = parser.parse_args()
    
    generator = TestDataGenerator(output_dir=args.output_dir)
    
    if args.custom:
        # 사용자 정의 파일 생성
        generator.create_xlsx_file(
            filename=args.filename,
            num_sheets=args.sheets,
            rows_per_sheet=args.rows,
            columns_per_sheet=args.columns,
            add_formatting=args.with_formatting,
            verbose=True
        )
    else:
        # 표준 테스트 스위트 생성
        generator.generate_test_suite()


if __name__ == "__main__":
    main()
